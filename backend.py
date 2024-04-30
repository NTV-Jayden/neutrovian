import openpyxl
import pandas as pd
import csv
from datetime import datetime
import os


def gather_product_lot(lot_file):
    '''
        prod_lot = {"prod_1": True}
    '''
    prod_lot = {}
    lot_workbook = openpyxl.load_workbook(lot_file, data_only=True)
    lot_sheet = lot_workbook.active
    max_row = lot_sheet.max_row
    for i in range(2, max_row+1):
        product = lot_sheet.cell(row=i, column=1).value
        tracking = lot_sheet.cell(row=i, column=2).value
        prod_lot[product] = True if tracking == "By Lots" else False
    return prod_lot


def gather_wms_stock(wms_file, lot_file) -> dict:
        '''
            wms_stock = {('prod1', 'location', 'lot_number', 'uom'): quantity}
        '''
        wms_stock = {}
        wms_workbook = openpyxl.load_workbook(wms_file, data_only=True)
        wms_sheet = wms_workbook.active
        max_row = wms_sheet.max_row
        product_lot = gather_product_lot(lot_file)
        for i in range(2, max_row+1):
            product = wms_sheet.cell(row=i, column=3).value
            quantity = wms_sheet.cell(row=i, column=12).value
            tracking = product_lot[product]
            lot_number = ""
            if tracking:
                lot_number = str(wms_sheet.cell(row=i, column=6).value)
            location = wms_sheet.cell(row=i, column=1).value + '/Stock'
            uom = wms_sheet.cell(row=i, column=13).value
            product_key = (product, location, lot_number, uom.upper())
            if product_key not in wms_stock:
                if quantity:
                    wms_stock[product_key] = quantity
            else:
                wms_stock[product_key] += quantity
        return wms_stock


def gather_odoo_stock(odoo_file) -> dict:
    '''
        odoo_stock = {('prod1', 'location', 'lot_number', 'uom'): (quantity, 'lot_external_id')}
    '''
    odoo_stock = {}
    odoo_workbook = openpyxl.load_workbook(odoo_file, data_only=True)
    odoo_sheet = odoo_workbook.active
    max_row = odoo_sheet.max_row
    for i in range(2, max_row+1):
        product = odoo_sheet.cell(row=i, column=4).value
        tracking = odoo_sheet.cell(row=i, column=2).value
        quantity = odoo_sheet.cell(row=i, column=5).value
        lot_number = ""
        lot_external_id = ""
        if tracking == 'lot':
            lot_number = odoo_sheet.cell(row=i, column=3).value
            lot_external_id = odoo_sheet.cell(row=i, column=7).value
        location = odoo_sheet.cell(row=i, column=1).value
        uom = odoo_sheet.cell(row=i, column=6).value
        product_key = (product, location, lot_number, uom.upper())
        if product_key not in odoo_stock:
            if quantity:
                odoo_stock[product_key] = (quantity, lot_external_id)
        else:
            odoo_stock[product_key] = (quantity + odoo_stock[product_key][0], lot_external_id)
    return odoo_stock


def compare_stock(odoo: dict, wms: dict) -> dict:
    '''
        adjustment = {('prod1', 'location', 'lot_number', 'uom'): (quantity, 'lot_external_id', 'reason')}
    '''        
    adjustment = {}

    for product_key, wms_quantity in wms.items():
        if product_key in odoo:
            if wms_quantity != odoo[product_key][0]:
                adjustment[product_key] = (wms_quantity, odoo[product_key][0], odoo[product_key][1], 'odoo & wms different')
        elif product_key not in odoo:
            adjustment[product_key] = (wms_quantity, 0, "Find LNEI", 'wms qty not found in odoo')

    for product_key, odoo_quantity in odoo.items():
        if product_key not in wms:
            adjustment[product_key] = (0, odoo_quantity[0], "Find LNEI", 'odoo qty not found in wms')
    return adjustment


def write_result(adjustment: dict[tuple]) -> None:
    current_datetime = datetime.now()
    csv_filename = current_datetime.strftime("%Y%m%d_%H%M%S") + ".csv"
    excel_filename = "difference_on_" + current_datetime.strftime("%Y%m%d_%H%M%S") + ".xlsx"

    with open(csv_filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Location', 'Product', 'WMS Qty', 'Odoo Qty', 'Adjustment', 'UOM', 'Lot Number', 'Lot External ID', 'Reason'])
        for product_key, quantity in adjustment.items():
            try:
                product, location, lot_number, uom = product_key
                wms_qty, odoo_qty, lot_external_id, reason = quantity
                writer.writerow([location, product, wms_qty, odoo_qty, wms_qty-odoo_qty, uom, lot_number, lot_external_id, reason])
            except ValueError:
                print(product_key)
    df = pd.read_csv(csv_filename)
    df.to_excel(excel_filename, index=False)
    os.remove(csv_filename)
    return excel_filename


def run_comparison(odoo_file, wms_file, lot_file):
    wms_stock = gather_wms_stock(wms_file, lot_file)
    odoo_stock = gather_odoo_stock(odoo_file)
    adjustment = compare_stock(odoo_stock, wms_stock)
    return write_result(adjustment)
